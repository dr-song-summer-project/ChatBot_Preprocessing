from collections import defaultdict
import nltk
import pandas as pd
from soynlp.normalizer import *
import numpy as np
import similarity
from openpyxl import Workbook

MAXSIZE = 10

path_sample = 'src/sample_short.xlsx'
# path_training = 'src/categories.xlsx'
test_path = 'src/test2.xlsx'  # 카테고리 항목

def read_file(file_paths):  # ==파일 읽어오기
    df_ = pd.read_excel(file_paths)
    numpy = pd.DataFrame.to_numpy(df_)
    return df_, numpy


def classify(list_, text_Q, text_A):
    correct = True

    print('\n')
    print('질문 = '+text_Q)
    print('답변 = '+text_A)

    ant_Q = input("질문 요약문을 입력하세요 : ")
    ant_A = input("답변 요약문을 입력하세요 : ")
    while correct:
        select = (input('0~%d중 분류를 입력하세요 : ' % (len(list_) - 1)).split())
        for j in range(len(select)):
            if int(select[j]) < len(list_) and select[j] != '':
                correct = False
            else:
                correct = True

    Idx.append(" ".join(select))
    Query.append((text_Q, ant_Q))
    Answer.append((text_A, ant_A))


# =============================================================

text_df, text_df_np = read_file(path_sample)  # 학습할 데이터
ct_df, ct_df_np = read_file(test_path)  # 데이터를 나눌 카테고리

# =============================================================

data = pd.DataFrame()
standard = pd.DataFrame()
data['querys'] = text_df['질문']
data['answers'] = text_df['답변']
standard['category3'] = ct_df['Category4']
# display(standard['category3'])

standard = standard[standard['category3'].notna()]
data = data[data['querys'].notna()]
data = data[data['answers'].notna()]
# category3 = [[0 for j in range(MAXSIZE)] for i in range(len(standard))]
category3 = []

# =============================================================

tmp = standard['category3'].values.tolist()
Qst = data['querys'].values.tolist()
Ans = data['answers'].values.tolist()

# =============================================================

Idx = []
Query = []
Answer = []

# =============================================================

for i in range(len(data['querys'])):
    # category 출력
    for j in range(len(standard['category3'])):
        # category3[j] = tmp[j]
        print(j, '번: ', tmp[j], end=' ')
        if j % 6 == 0:
            print()
    # 코사인 유사도 측정 카테고리 출력
    # a = s.get_recommendation(str(Qst[i]))
    # 분류
    classify(tmp, Qst[i], Ans[i])

# ============================================
# --  Excel 입력
# ============================================
wb = Workbook()
sheet = wb.active
sheet.title = '분류된 문장' # 컬럼명 지정(헤더)

sheet.cell(row=1, column=1, value='인덱스')
sheet.cell(row=1, column=2, value='질문')
sheet.cell(row=1, column=3, value='질문 요약')
sheet.cell(row=1, column=4, value='답변')
sheet.cell(row=1, column=5, value='답변 요약')


# 시트 저장
row_no = 2
for n, rows in enumerate(Query):
  for seq, value in enumerate(rows):
    sheet.cell(row=row_no+n, column=1, value=Idx[n])
    sheet.cell(row=row_no+n, column=2, value=Query[n][0])
    sheet.cell(row=row_no+n, column=3, value=Query[n][1])
    sheet.cell(row=row_no+n, column=4, value=Answer[n][0])
    sheet.cell(row=row_no+n, column=5, value=Answer[n][1])

wb.save('src/sample_classified.xlsx')
wb.close()